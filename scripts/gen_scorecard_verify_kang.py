# -*- coding: utf-8 -*-
"""강환국 스코어카드 검증 엑셀 생성"""
import sys, json, os
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(__file__))
from pipeline_config import PipelineConfig
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import requests

config = PipelineConfig()
headers = {'apikey': config.SUPABASE_SERVICE_KEY, 'Authorization': f'Bearer {config.SUPABASE_SERVICE_KEY}'}

# Load scorecard
sc = json.load(open(os.path.join(os.path.dirname(__file__), '..', 'data', 'influencer_scorecard.json'), encoding='utf-8'))
kang = sc['speakers']['kang-hwanguk']
scored_map = {s['id']: s for s in kang['scored_list']}

# Load stock prices
stock_prices = json.load(open(os.path.join(os.path.dirname(__file__), '..', 'data', 'stockPrices.json'), encoding='utf-8'))

# Get 129 signals
speaker_id = '4ab064b4-79ee-4da8-bc47-c4feeb9f7b73'
r = requests.get(
    config.SUPABASE_URL + '/rest/v1/influencer_signals?speaker_id=eq.' + speaker_id +
    '&select=id,stock,ticker,signal,key_quote,reasoning,video_id,influencer_videos(title,video_id,published_at)&limit=200',
    headers=headers
)
all_signals = r.json()
print(f'Total: {len(all_signals)}건')

# Return calc
def find_closest(prices, target_str, max_days=30):
    target = datetime.strptime(target_str, '%Y-%m-%d')
    best, best_diff = None, 999
    for p in prices:
        try:
            d = datetime.strptime(p['date'], '%Y-%m-%d')
        except:
            continue
        diff = abs((d - target).days)
        if diff < best_diff and diff <= max_days:
            best = p['close']
            best_diff = diff
    return best

def get_prices(ticker):
    v = stock_prices.get(ticker)
    if v is None and ticker and not ticker.endswith('.KS'):
        v = stock_prices.get(ticker + '.KS')
    if v is None:
        return []
    return v.get('prices', []) if isinstance(v, dict) else (v if isinstance(v, list) else [])

def calc_returns(ticker, date_str):
    prices = get_prices(ticker)
    if not prices or not date_str:
        return None, None, None
    price_at = find_closest(prices, date_str)
    if not price_at:
        return None, None, None
    price_latest = prices[-1]['close']
    r_cur = round((price_latest - price_at) / price_at * 100, 2) if price_latest else None
    try:
        sd = datetime.strptime(date_str, '%Y-%m-%d')
    except:
        return None, None, r_cur
    t1y = sd + timedelta(days=365)
    r_1y = None
    if t1y <= datetime.now():
        p1y = find_closest(prices, t1y.strftime('%Y-%m-%d'))
        r_1y = round((p1y - price_at) / price_at * 100, 2) if p1y else None
    t3y = sd + timedelta(days=1095)
    r_3y = None
    if t3y <= datetime.now():
        p3y = find_closest(prices, t3y.strftime('%Y-%m-%d'))
        r_3y = round((p3y - price_at) / price_at * 100, 2) if p3y else None
    return r_1y, r_3y, r_cur

now = datetime.now()
EVAL_GRACE_DAYS = 90
BUY_SIGNALS = {'매수', '긍정', 'STRONG_BUY', 'BUY', 'POSITIVE'}
SELL_SIGNALS = {'매도', 'SELL', 'STRONG_SELL'}
HIT_SIGNALS = BUY_SIGNALS | SELL_SIGNALS

# Map DB signals by id for key_quote/title/url
db_map = {}
for sig in all_signals:
    vid = sig.get('influencer_videos') or {}
    yt_id = vid.get('video_id', '')
    db_map[sig['id']] = {
        'key_quote': sig.get('key_quote', ''),
        'reasoning': sig.get('reasoning', ''),
        'title': vid.get('title', ''),
        'url': f'https://www.youtube.com/watch?v={yt_id}' if yt_id else '',
        'pub': (vid.get('published_at', '') or '')[:10],
    }

# Build enriched list from scored_list (96건, dedup 후) — 스코어카드와 정확히 일치
enriched = []
for s in kang['scored_list']:
    sid = s['id']
    db = db_map.get(sid, {})
    r_1y = s.get('return_1y')
    r_3y = s.get('return_3y')
    r_cur = s.get('return_current')
    date = s.get('date', '')
    signal = s.get('signal', '')

    try:
        days = (now - datetime.strptime(date, '%Y-%m-%d')).days
    except:
        days = 0

    # W/L from scored_list hit field
    hit = s.get('hit')
    if hit is True:
        wl = 'W'
    elif hit is False:
        wl = 'L'
    elif signal in HIT_SIGNALS:
        wl = 'pending'
    else:
        wl = 'N/A'

    # mid return
    mid_ret = r_3y if (days >= 1095 and r_3y is not None) else r_cur

    enriched.append({
        'stock': s.get('stock', ''), 'ticker': s.get('ticker', ''),
        'signal': signal, 'key_quote': db.get('key_quote', ''),
        'reasoning': db.get('reasoning', ''), 'title': db.get('title', ''),
        'url': db.get('url', ''), 'date': date,
        'r_1y': r_1y, 'r_3y': r_3y, 'r_cur': r_cur, 'mid_ret': mid_ret,
        'wl': wl, 'days': days, 'is_hit': signal in HIT_SIGNALS,
    })

enriched.sort(key=lambda x: x['date'])

# Tier classification matching calc_influencer_scorecard.py exactly
swing, mid, long_ = [], [], []
for e in enriched:
    if not e['is_hit']:
        continue
    days = e['days']
    r_1y, r_3y, r_cur = e['r_1y'], e['r_3y'], e['r_cur']

    # swing: min_days=90, ret_key=return_1y, fallback current
    if days >= 90:
        ret = r_1y if r_1y is not None else (r_cur if days >= EVAL_GRACE_DAYS else None)
        if ret is not None:
            swing.append(e)

    # mid: min_days=365
    if days >= 365:
        if days >= 1095 and r_3y is not None:
            mid.append(e)
        elif r_cur is not None:
            mid.append(e)

    # long: min_days=1095, ret_key=return_current
    if days >= 1095:
        ret = r_cur if r_cur is not None else None
        if ret is not None:
            long_.append(e)

print(f'스윙: {len(swing)}, 중기: {len(mid)}, 장기: {len(long_)}')

# Excel
wb = Workbook()
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True, size=10)
win_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
loss_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

def write_sheet(ws, title, items, ret_name, ret_key):
    ws.title = title
    cols = ['번호', '종목명', 'ticker', '날짜', 'sentiment', '발언원문', '영상제목', ret_name, 'W/L']
    widths = [5, 18, 10, 12, 8, 55, 45, 12, 6]
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center'); cell.border = thin
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for idx, item in enumerate(items, 1):
        row = idx + 1
        ret_val = item.get(ret_key)
        vals = [idx, item['stock'], item['ticker'], item['date'], item['signal'],
                item['key_quote'] or item['reasoning'], item['title'],
                f"{ret_val:+.1f}%" if ret_val is not None else '', item['wl']]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = thin
            cell.alignment = Alignment(vertical='top', wrap_text=(c in (6, 7)))
        wl_cell = ws.cell(row=row, column=9)
        if item['wl'] == 'W': wl_cell.fill = win_fill
        elif item['wl'] == 'L': wl_cell.fill = loss_fill
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:I{len(items)+1}'

ws1 = wb.active
write_sheet(ws1, f'스윙 ({len(swing)}건)', swing, '1Y 수익률', 'r_1y')

ws2 = wb.create_sheet()
write_sheet(ws2, f'중기 ({len(mid)}건)', mid, '1~3Y 수익률', 'mid_ret')

ws3 = wb.create_sheet()
write_sheet(ws3, f'장기 ({len(long_)}건)', long_, '3Y+ 수익률', 'r_cur')

# Sheet 4: 로직 설명
ws4 = wb.create_sheet(title='분류기준 + 수익률 로직')
ws4.column_dimensions['A'].width = 85

swing_t = kang['tiers']['swing']
mid_t = kang['tiers']['mid']
long_t = kang['tiers']['long']

lines = [
    '=== 스윙/중기/장기 분류 기준 ===',
    '',
    '분류는 time_horizon 필드가 아니라 "시그널 날짜로부터 경과 일수"로 결정',
    '파일: scripts/calc_influencer_scorecard.py (lines 471-474)',
    '',
    '  스윙 (swing): 시그널 날짜 >= 90일 경과. return_1y (1년후 수익률) 사용',
    '    - calc_tier_simple(scored, "return_1y", min_days=90)',
    '    - 1Y 없으면 90일+ 경과 시 return_current fallback',
    '',
    '  중기 (mid): 시그널 날짜 >= 365일 경과',
    '    - calc_mid_tier(scored):',
    '      1095일+ 경과 & return_3y 존재 -> return_3y 사용',
    '      그 외 -> return_current (현재가 기준) 사용',
    '',
    '  장기 (long): 시그널 날짜 >= 1095일(3년) 경과. return_current 사용',
    '    - calc_tier_simple(scored, "return_current", min_days=1095)',
    '',
    '대상 시그널: 매수, 긍정, 매도만 (중립/부정 제외)',
    '  BUY_SIGNALS = {매수, 긍정}',
    '  SELL_SIGNALS = {매도}',
    '  부정(CONCERN)은 적중률 계산에서 제외',
    '',
    '월별 중복 제거: (speaker + ticker + YYYY-MM) 기준 1건만 유지',
    '',
    '',
    '=== 수익률 계산 로직 ===',
    '',
    '파일: scripts/calc_influencer_scorecard.py (lines 79-126)',
    '가격: data/stockPrices.json',
    '',
    '  수익률 = (비교가 - 시그널시점가) / 시그널시점가 x 100',
    '',
    '  1. price_at_signal: 시그널 날짜에서 +-30일 내 가장 가까운 종가',
    '  2. return_1y: 시그널 날짜 + 365일 시점 종가 기준',
    '  3. return_3y: 시그널 날짜 + 1095일 시점 종가 기준',
    '  4. return_current: 가장 최근 종가 기준',
    '',
    'W/L 판정:',
    '  매수/긍정: 수익률 > 0 -> W, <= 0 -> L',
    '  매도:     수익률 < 0 -> W, >= 0 -> L',
    '',
    '하이브리드 판정 (lines 254-263):',
    '  1순위: return_1y',
    '  2순위: 90일+ 경과 시 return_current',
    '  그 외: pending',
    '',
    'hit_rate = wins / (wins + losses) x 100',
    '  tier별 최소 5건, 전체 최소 10건',
    '',
    '',
    '=== 강환국 스코어카드 결과 ===',
    '',
    f'  스윙: {swing_t["count"]}건, 적중률 {swing_t["hit_rate"]}%, W {swing_t["wins"]} / L {swing_t["losses"]}, 중앙수익률 {swing_t["median_return"]}%',
    f'  중기: {mid_t["count"]}건, 적중률 {mid_t["hit_rate"]}%, W {mid_t["wins"]} / L {mid_t["losses"]}, 중앙수익률 {mid_t["median_return"]}%',
    f'  장기: {long_t["count"]}건, 적중률 {long_t["hit_rate"]}%, W {long_t["wins"]} / L {long_t["losses"]}, 중앙수익률 {long_t["median_return"]}%',
]

for i, line in enumerate(lines, 1):
    cell = ws4.cell(row=i, column=1, value=line)
    if line.startswith('==='):
        cell.font = Font(bold=True, size=12)
    elif line.startswith('  ') and not line.startswith('    '):
        cell.font = Font(size=10)

out = os.path.join(os.path.dirname(__file__), '..', 'data', 'scorecard_verify_kang.xlsx')
wb.save(out)
print(f'저장: {out}')
