# -*- coding: utf-8 -*-
"""eval_excel_26.py - V11.5 vs V12 eval 26개 + 엑셀 저장"""
import sys, os, json, time, re, requests
sys.stdout.reconfigure(encoding='utf-8')
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
load_dotenv(ROOT / '.env.local')
import anthropic
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

URL  = os.environ['NEXT_PUBLIC_SUPABASE_URL']
KEY  = os.environ.get('SUPABASE_SERVICE_ROLE_KEY') or os.environ['NEXT_PUBLIC_SUPABASE_ANON_KEY']
AKEY = os.environ['ANTHROPIC_API_KEY']
H    = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Accept': 'application/json'}

MODEL       = 'claude-haiku-4-5-20251001'
VALID       = {'매수','긍정','중립','부정','매도'}
PROMPT_V115 = ROOT / 'prompts' / 'pipeline_v11.5.md'
PROMPT_V12  = ROOT / 'prompts' / 'pipeline_v12.md'

TS = datetime.now().strftime('%Y%m%d_%H%M')
XLSX_PATH = ROOT / 'data' / f'eval_v11v12_{TS}.xlsx'
JSON_PATH = ROOT / 'data' / f'eval_v11v12_{TS}.json'

def log(msg): print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}', flush=True)

# ─── 색상 정의 ───────────────────────────────────────────────────────────────
CLR = {
    'header_blue':   PatternFill('solid', fgColor='1F4E79'),
    'header_green':  PatternFill('solid', fgColor='375623'),
    'header_orange': PatternFill('solid', fgColor='843C0C'),
    'header_gray':   PatternFill('solid', fgColor='404040'),
    'correct':       PatternFill('solid', fgColor='C6EFCE'),
    'wrong':         PatternFill('solid', fgColor='FFC7CE'),
    'over':          PatternFill('solid', fgColor='FFEB9C'),
    'under':         PatternFill('solid', fgColor='BDD7EE'),
    'row_even':      PatternFill('solid', fgColor='F2F2F2'),
    'row_odd':       PatternFill('solid', fgColor='FFFFFF'),
    'sig_매수':      PatternFill('solid', fgColor='DAEEF3'),
    'sig_긍정':      PatternFill('solid', fgColor='EBF1DE'),
    'sig_중립':      PatternFill('solid', fgColor='F2F2F2'),
    'sig_부정':      PatternFill('solid', fgColor='FDE9D9'),
    'sig_매도':      PatternFill('solid', fgColor='FFC7CE'),
}
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

def hdr(ws, row, col, val, fill_key='header_blue', bold=True, color='FFFFFF', wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = CLR[fill_key]
    c.font = Font(bold=bold, color=color, size=10)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    c.border = THIN
    return c

def cell(ws, row, col, val, fill=None, bold=False, align='left', color='000000'):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    c.font = Font(bold=bold, color=color, size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = THIN
    return c

# ─── DB ────────────────────────────────────────────────────────────────────────
def get_videos():
    r = requests.get(
        f'{URL}/rest/v1/influencer_videos'
        f'?select=id,video_id,title,duration_seconds,channel_id,subtitle_text'
        f'&signal_count=gt.0&subtitle_text=not.is.null'
        f'&order=published_at.desc&limit=26', headers=H)
    videos = r.json()
    ch_r = requests.get(f'{URL}/rest/v1/influencer_channels?select=id,channel_name', headers=H)
    ch   = {c['id']: c['channel_name'] for c in ch_r.json()}
    for v in videos:
        v['channel_name'] = ch.get(v['channel_id'], '?')
    return videos

def get_gt(db_id):
    r = requests.get(
        f'{URL}/rest/v1/influencer_signals?select=stock,signal&video_id=eq.{db_id}', headers=H)
    data = r.json()
    if not isinstance(data, list): return []
    return [{'stock': s['stock'], 'signal_type': s['signal']} for s in data if isinstance(s, dict)]

# ─── 프롬프트 ──────────────────────────────────────────────────────────────────
def build_prompt(path, video):
    tmpl = path.read_text(encoding='utf-8')
    sub  = video.get('subtitle_text', '')
    if len(sub) > 40000: sub = sub[:40000] + '\n...(생략)'
    dur  = video.get('duration_seconds', 0)
    dur_str = f'{dur//60}분 {dur%60}초' if dur else '알 수 없음'
    header = (f'[EVAL MODE - 자막 기반]\n영상 제목: {video.get("title","")}\n'
              f'영상 길이: {dur_str}\n채널: {video.get("channel_name","")}\n\n'
              f'=== 자막 ===\n{sub}\n=== 자막 끝 ===\n\n위 자막을 분석하세요:\n\n')
    p = header + tmpl
    p = p.replace('{VIDEO_DURATION_INFO}', dur_str).replace('{CHANNEL_URL}', '')
    return p

# ─── Claude 호출 ───────────────────────────────────────────────────────────────
def call_claude(prompt, label=''):
    client = anthropic.Anthropic(api_key=AKEY)
    for attempt in range(1, 4):
        try:
            msg = client.messages.create(
                model=MODEL, max_tokens=4096,
                messages=[{'role':'user','content': prompt + '\n\nJSON만: {"signals":[...]}'}])
            raw = msg.content[0].text.strip()
            for pat in [r'```json\s*(.*?)\s*```', r'(\{.*\})', None]:
                try:
                    text = re.search(pat, raw, re.DOTALL).group(1) if pat else raw
                    sigs = json.loads(text).get('signals', [])
                    return [s for s in sigs if isinstance(s,dict) and s.get('signal_type') in VALID]
                except: pass
            return []
        except anthropic.RateLimitError:
            log(f'[RATE] 30초 대기'); time.sleep(30)
        except Exception as e:
            log(f'[ERR] {e}')
            if attempt < 3: time.sleep(5)
    return []

# ─── 비교 ─────────────────────────────────────────────────────────────────────
def compare(gt, pred):
    gm = {s['stock'].strip(): s['signal_type'] for s in gt}
    pm = {s['stock'].strip(): s.get('signal_type','') for s in pred}
    correct, wrong, over, under = 0, [], [], []
    for st, gs in gm.items():
        if st in pm:
            if pm[st] == gs: correct += 1
            else: wrong.append({'stock':st,'gt':gs,'pred':pm[st]})
        else: under.append({'stock':st,'gt':gs})
    for st, ps in pm.items():
        if st not in gm: over.append({'stock':st,'pred':ps})
    return dict(correct=correct, total=len(gm), acc=correct/len(gm) if gm else None,
                wrong=wrong, over=over, under=under)

# ─── 엑셀 생성 ────────────────────────────────────────────────────────────────
def make_excel(results):
    wb = Workbook()

    # ── Sheet1: 요약 ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = '요약'
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    hdr(ws,1,1,'항목', 'header_gray')
    hdr(ws,1,2,'V11.5', 'header_blue')
    hdr(ws,1,3,'V12', 'header_green')
    hdr(ws,1,4,'개선', 'header_orange')

    gt_tot = sum(r['v115_cmp']['total'] for r in results)
    c115   = sum(r['v115_cmp']['correct'] for r in results)
    c12    = sum(r['v12_cmp']['correct'] for r in results)
    w115   = sum(len(r['v115_cmp']['wrong']) for r in results)
    w12    = sum(len(r['v12_cmp']['wrong']) for r in results)
    o115   = sum(len(r['v115_cmp']['over']) for r in results)
    o12    = sum(len(r['v12_cmp']['over']) for r in results)
    u115   = sum(len(r['v115_cmp']['under']) for r in results)
    u12    = sum(len(r['v12_cmp']['under']) for r in results)
    a115   = c115/gt_tot if gt_tot else 0
    a12    = c12/gt_tot  if gt_tot else 0

    rows = [
        ('평가 영상 수', len(results), len(results), '—'),
        ('총 GT 시그널 수', gt_tot, gt_tot, '—'),
        ('정확 일치 수', c115, c12, f'+{c12-c115}' if c12>=c115 else str(c12-c115)),
        ('정확도 (signal_type 일치율)', f'{a115:.1%}', f'{a12:.1%}',
         f'{"+" if a12>=a115 else ""}{(a12-a115):.1%}'),
        ('오분류 수', w115, w12, f'+{w12-w115}' if w12>=w115 else str(w12-w115)),
        ('과잉추출 수 (GT에 없는 종목)', o115, o12, f'+{o12-o115}' if o12>=o115 else str(o12-o115)),
        ('미추출 수 (GT에 있으나 미탐지)', u115, u12, f'+{u12-u115}' if u12>=u115 else str(u12-u115)),
    ]
    for i, (label, v1, v2, diff) in enumerate(rows, 2):
        fill = CLR['row_even'] if i%2==0 else CLR['row_odd']
        cell(ws, i, 1, label, fill, bold=True)
        cell(ws, i, 2, v1, fill, align='center')
        cell(ws, i, 3, v2, fill, align='center')
        is_imp = str(diff).startswith('+') and diff != '+0' and diff != '+0.0%'
        is_bad = not str(diff).startswith('+') and diff not in ('—', '+0', '+0.0%')
        diff_fill = CLR['correct'] if is_imp else CLR['wrong'] if is_bad else fill
        cell(ws, i, 4, diff, diff_fill, align='center')

    ws.row_dimensions[1].height = 22
    for i in range(2, len(rows)+2):
        ws.row_dimensions[i].height = 20

    # ── Sheet2: 영상별 상세 ───────────────────────────────────────────────────
    ws2 = wb.create_sheet('영상별_상세')
    cols = ['#','영상ID','채널','제목','GT 종목','GT 시그널',
            'V11.5 종목','V11.5 시그널','V11.5 conf',
            'V12 종목', 'V12 시그널',  'V12 conf',
            'V11.5 결과','V12 결과']
    widths = [4,14,18,45,14,10,14,10,8,14,10,8,10,10]
    for i,(c,w) in enumerate(zip(cols,widths),1):
        hdr(ws2,1,i,c,'header_gray')
        ws2.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for idx, r in enumerate(results, 1):
        v    = r['video']
        gt   = r['video']['gt']
        s115 = r['v115_sigs']
        s12  = r['v12_sigs']
        gt_map  = {x['stock']: x['signal_type'] for x in gt}
        p115_map = {x['stock']: (x.get('signal_type',''), x.get('confidence','')) for x in s115}
        p12_map  = {x['stock']: (x.get('signal_type',''), x.get('confidence','')) for x in s12}

        all_stocks = (list(gt_map.keys()) +
                      [s for s in p115_map if s not in gt_map] +
                      [s for s in p12_map if s not in gt_map and s not in p115_map])

        row_fill = CLR['row_even'] if idx%2==0 else CLR['row_odd']
        for si, stock in enumerate(all_stocks):
            gt_sig   = gt_map.get(stock, '')
            s115_sig, s115_c = p115_map.get(stock, ('',''))
            s12_sig,  s12_c  = p12_map.get(stock, ('',''))

            # 결과 판정
            res115 = ('✅ 정확' if s115_sig == gt_sig and gt_sig else
                      '❌ 오분류' if gt_sig and s115_sig else
                      '⚠ 과잉' if not gt_sig and s115_sig else
                      '📭 미탐지' if gt_sig and not s115_sig else '')
            res12  = ('✅ 정확' if s12_sig == gt_sig and gt_sig else
                      '❌ 오분류' if gt_sig and s12_sig else
                      '⚠ 과잉' if not gt_sig and s12_sig else
                      '📭 미탐지' if gt_sig and not s12_sig else '')

            fill115 = (CLR['correct'] if '정확' in res115 else
                       CLR['wrong'] if '오분류' in res115 else
                       CLR['over'] if '과잉' in res115 else
                       CLR['under'] if '미탐지' in res115 else row_fill)
            fill12  = (CLR['correct'] if '정확' in res12 else
                       CLR['wrong'] if '오분류' in res12 else
                       CLR['over'] if '과잉' in res12 else
                       CLR['under'] if '미탐지' in res12 else row_fill)

            cell(ws2, row, 1,  idx if si==0 else '', row_fill, align='center')
            cell(ws2, row, 2,  v['video_id'] if si==0 else '', row_fill)
            cell(ws2, row, 3,  v['channel_name'] if si==0 else '', row_fill)
            cell(ws2, row, 4,  v['title'][:60] if si==0 else '', row_fill)
            cell(ws2, row, 5,  stock, row_fill, bold=bool(gt_sig))
            sig_fill = CLR.get(f'sig_{gt_sig}', row_fill)
            cell(ws2, row, 6,  gt_sig, sig_fill, align='center')
            cell(ws2, row, 7,  stock if s115_sig else '', row_fill)
            s115_fill = CLR.get(f'sig_{s115_sig}', row_fill)
            cell(ws2, row, 8,  s115_sig, s115_fill, align='center')
            cell(ws2, row, 9,  s115_c, row_fill, align='center')
            cell(ws2, row, 10, stock if s12_sig else '', row_fill)
            s12_fill = CLR.get(f'sig_{s12_sig}', row_fill)
            cell(ws2, row, 11, s12_sig, s12_fill, align='center')
            cell(ws2, row, 12, s12_c, row_fill, align='center')
            cell(ws2, row, 13, res115, fill115, align='center')
            cell(ws2, row, 14, res12,  fill12,  align='center')
            ws2.row_dimensions[row].height = 18
            row += 1

    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = 'A2'

    # ── Sheet3: 오분류 상세 ───────────────────────────────────────────────────
    ws3 = wb.create_sheet('오분류_상세')
    hdr(ws3,1,1,'영상ID','header_gray'); ws3.column_dimensions['A'].width=14
    hdr(ws3,1,2,'채널','header_gray');   ws3.column_dimensions['B'].width=18
    hdr(ws3,1,3,'제목','header_gray');   ws3.column_dimensions['C'].width=40
    hdr(ws3,1,4,'종목','header_gray');   ws3.column_dimensions['D'].width=14
    hdr(ws3,1,5,'GT','header_gray');     ws3.column_dimensions['E'].width=8
    hdr(ws3,1,6,'V11.5 예측','header_blue'); ws3.column_dimensions['F'].width=12
    hdr(ws3,1,7,'V12 예측','header_green');  ws3.column_dimensions['G'].width=12
    hdr(ws3,1,8,'V11.5 판정','header_blue'); ws3.column_dimensions['H'].width=12
    hdr(ws3,1,9,'V12 판정','header_green');  ws3.column_dimensions['I'].width=12

    # 모든 오분류/미탐지 종목 수집
    all_issues = []
    for r in results:
        v = r['video']
        gt_map   = {x['stock']: x['signal_type'] for x in v['gt']}
        p115_map = {x['stock']: x.get('signal_type','') for x in r['v115_sigs']}
        p12_map  = {x['stock']: x.get('signal_type','') for x in r['v12_sigs']}
        all_stocks = set(gt_map)|set(p115_map)|set(p12_map)
        for stock in all_stocks:
            gs  = gt_map.get(stock,'')
            p115 = p115_map.get(stock,'')
            p12  = p12_map.get(stock,'')
            judge115 = ('정확' if gs and p115==gs else '오분류' if gs and p115 and p115!=gs
                        else '과잉' if not gs and p115 else '미탐지' if gs and not p115 else '')
            judge12  = ('정확' if gs and p12==gs else '오분류' if gs and p12 and p12!=gs
                        else '과잉' if not gs and p12 else '미탐지' if gs and not p12 else '')
            if '정확' not in judge115 or '정확' not in judge12:
                all_issues.append((v, stock, gs, p115, p12, judge115, judge12))

    for i, (v, stock, gs, p115, p12, j115, j12) in enumerate(all_issues, 2):
        rf = CLR['row_even'] if i%2==0 else CLR['row_odd']
        f115 = (CLR['correct'] if j115=='정확' else CLR['wrong'] if j115=='오분류'
                else CLR['over'] if j115=='과잉' else CLR['under'] if j115=='미탐지' else rf)
        f12  = (CLR['correct'] if j12=='정확'  else CLR['wrong'] if j12=='오분류'
                else CLR['over'] if j12=='과잉'  else CLR['under'] if j12=='미탐지' else rf)
        cell(ws3,i,1,v['video_id'],rf)
        cell(ws3,i,2,v['channel_name'],rf)
        cell(ws3,i,3,v['title'][:50],rf)
        cell(ws3,i,4,stock,rf,bold=True)
        cell(ws3,i,5,gs,CLR.get(f'sig_{gs}',rf),align='center')
        cell(ws3,i,6,p115,CLR.get(f'sig_{p115}',rf),align='center')
        cell(ws3,i,7,p12, CLR.get(f'sig_{p12}',rf), align='center')
        cell(ws3,i,8,j115,f115,align='center')
        cell(ws3,i,9,j12, f12, align='center')
        ws3.row_dimensions[i].height = 18
    ws3.row_dimensions[1].height = 22
    ws3.freeze_panes = 'A2'

    # ── Sheet4: 오분류 패턴 분석 ─────────────────────────────────────────────
    ws4 = wb.create_sheet('오분류_패턴')
    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 10
    ws4.column_dimensions['D'].width = 10

    hdr(ws4,1,1,'GT','header_gray')
    hdr(ws4,1,2,'예측','header_gray')
    hdr(ws4,1,3,'V11.5 건수','header_blue')
    hdr(ws4,1,4,'V12 건수','header_green')

    from collections import Counter
    pat115 = Counter()
    pat12  = Counter()
    for r in results:
        for w in r['v115_cmp']['wrong']: pat115[(w['gt'],w['pred'])] += 1
        for w in r['v12_cmp']['wrong']:  pat12[(w['gt'],w['pred'])]  += 1
    all_pats = sorted(set(pat115)|set(pat12), key=lambda x: -(pat115.get(x,0)+pat12.get(x,0)))
    for i, (gt_s, pred_s) in enumerate(all_pats, 2):
        rf = CLR['row_even'] if i%2==0 else CLR['row_odd']
        cell(ws4,i,1,gt_s,  CLR.get(f'sig_{gt_s}',rf), align='center')
        cell(ws4,i,2,pred_s,CLR.get(f'sig_{pred_s}',rf),align='center')
        cell(ws4,i,3,pat115.get((gt_s,pred_s),0),CLR['wrong'] if pat115.get((gt_s,pred_s),0) else rf,align='center')
        cell(ws4,i,4,pat12.get((gt_s,pred_s),0), CLR['wrong'] if pat12.get((gt_s,pred_s),0)  else rf,align='center')
        ws4.row_dimensions[i].height = 18
    ws4.row_dimensions[1].height = 22

    wb.save(XLSX_PATH)
    log(f'엑셀 저장: {XLSX_PATH}')

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    log(f'=== eval_excel_26.py 시작 | 모델: {MODEL} ===')
    videos = get_videos()
    log(f'영상 수: {len(videos)}개')
    for v in videos:
        v['gt'] = get_gt(v['id'])

    results = []

    log('\n▶ V11.5 실행')
    for i, v in enumerate(videos, 1):
        log(f'  V11.5 [{i}/26] {v["video_id"]} | {v["title"][:35]}')
        t0 = time.time()
        sigs = call_claude(build_prompt(PROMPT_V115, v))
        cmp  = compare(v['gt'], sigs)
        log(f'    {time.time()-t0:.0f}초 | GT:{cmp["total"]} 정확:{cmp["correct"]} 추출:{len(sigs)}')
        results.append({'video':v,'v115_sigs':sigs,'v115_cmp':cmp,'v12_sigs':[],'v12_cmp':{}})
        if i%5==0 and i<26: log('  [배치] 10초 대기'); time.sleep(10)
        elif i<26: time.sleep(2)

    log('\n▶ V12 실행')
    for i, (v, r) in enumerate(zip(videos, results), 1):
        log(f'  V12  [{i}/26] {v["video_id"]} | {v["title"][:35]}')
        t0 = time.time()
        sigs = call_claude(build_prompt(PROMPT_V12, v))
        cmp  = compare(v['gt'], sigs)
        log(f'    {time.time()-t0:.0f}초 | GT:{cmp["total"]} 정확:{cmp["correct"]} 추출:{len(sigs)}')
        r['v12_sigs'] = sigs
        r['v12_cmp']  = cmp
        if i%5==0 and i<26: log('  [배치] 10초 대기'); time.sleep(10)
        elif i<26: time.sleep(2)

    # JSON 저장
    with open(JSON_PATH,'w',encoding='utf-8') as f:
        json.dump([{
            'video_id': r['video']['video_id'],
            'title':    r['video']['title'],
            'channel':  r['video']['channel_name'],
            'gt':       r['video']['gt'],
            'v115_sigs': r['v115_sigs'], 'v115_cmp': {k:v for k,v in r['v115_cmp'].items() if k!='acc'},
            'v12_sigs':  r['v12_sigs'],  'v12_cmp':  {k:v for k,v in r['v12_cmp'].items()  if k!='acc'},
        } for r in results], f, ensure_ascii=False, indent=2)
    log(f'JSON 저장: {JSON_PATH}')

    # 집계
    gt_tot = sum(r['v115_cmp']['total'] for r in results)
    c115   = sum(r['v115_cmp']['correct'] for r in results)
    c12    = sum(r['v12_cmp']['correct'] for r in results)
    a115   = c115/gt_tot if gt_tot else 0
    a12    = c12/gt_tot  if gt_tot else 0
    log(f'\n최종: V11.5={a115:.1%} ({c115}/{gt_tot}) | V12={a12:.1%} ({c12}/{gt_tot}) | 개선={"+" if a12>=a115 else ""}{(a12-a115):.1%}')

    make_excel(results)
    log('완료!')

if __name__ == '__main__':
    main()
