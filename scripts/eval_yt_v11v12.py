# -*- coding: utf-8 -*-
"""
eval_yt_v11v12.py
- eval_ground_truth_40.csv 기준 40행 평가
- yt-dlp Python API로 자막 추출 (캐시)
- V11.5 / V12.0 / V12.1 Haiku 분석
- 해당 stock 시그널만 비교 (다른 종목 무시)
- 결과: 마크다운 표 + 엑셀
"""
import sys, os, json, time, re, csv, subprocess
sys.stdout.reconfigure(encoding='utf-8')
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
load_dotenv(ROOT / '.env.local')

import anthropic
from youtube_transcript_api import YouTubeTranscriptApi, NoTranscriptFound, TranscriptsDisabled
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AKEY = os.environ['ANTHROPIC_API_KEY']
MODEL = 'claude-haiku-4-5-20251001'
VALID = {'매수','긍정','중립','부정','매도'}
PROMPT_V115 = ROOT / 'prompts' / 'pipeline_v11.5.md'
PROMPT_V120 = ROOT / 'prompts' / 'pipeline_v12.md'
PROMPT_V121 = ROOT / 'prompts' / 'pipeline_v12.1.md'
SUBTITLE_DIR = ROOT / 'data' / 'subtitles_cache'
SUBTITLE_DIR.mkdir(parents=True, exist_ok=True)

TS = datetime.now().strftime('%Y%m%d_%H%M')
RESULT_MD   = ROOT / 'data' / f'eval_yt_{TS}.md'
RESULT_XLSX = ROOT / 'data' / f'eval_yt_{TS}.xlsx'
PROGRESS    = ROOT / 'data' / f'eval_yt_{TS}_progress.json'

def log(msg): print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}', flush=True)

# ── 자막 추출 ─────────────────────────────────────────────────────────────────
def get_subtitle(video_id: str) -> str:
    cache_file = SUBTITLE_DIR / f'{video_id}.txt'
    if cache_file.exists():
        txt = cache_file.read_text(encoding='utf-8').strip()
        if txt:
            log(f'  [캐시] {video_id}')
            return txt

    log(f'  [transcript-api] {video_id} 추출 중...')
    api = YouTubeTranscriptApi()
    for langs in [['ko'], ['ko-KR'], ['en'], ['en-US'], None]:
        try:
            if langs:
                fetched = api.fetch(video_id, languages=langs)
            else:
                fetched = api.fetch(video_id)
            snippets = list(fetched)
            text = ' '.join(s.text for s in snippets)
            if text.strip():
                log(f'  [자막] {video_id} ({langs}) {len(text)}자')
                cache_file.write_text(text, encoding='utf-8')
                return text
        except (NoTranscriptFound, TranscriptsDisabled):
            continue
        except Exception as e:
            log(f'  [오류] {video_id}: {e}')
            continue

    log(f'  [자막없음] {video_id}')
    return ''

# ── 프롬프트 빌드 ─────────────────────────────────────────────────────────────
def build_prompt(prompt_path: Path, subtitle: str, video_id: str) -> str:
    tmpl = prompt_path.read_text(encoding='utf-8')
    sub = subtitle[:40000] + '\n...(생략)' if len(subtitle) > 40000 else subtitle
    header = (
        f'[EVAL MODE - 자막 기반]\n'
        f'영상 ID: {video_id}\n\n'
        f'=== 자막 ===\n{sub}\n=== 자막 끝 ===\n\n'
        f'위 자막을 분석하여 시그널을 추출하세요:\n\n'
    )
    p = header + tmpl
    p = p.replace('{VIDEO_DURATION_INFO}', '알 수 없음')
    p = p.replace('{CHANNEL_URL}', '')
    return p

# ── Claude 호출 ───────────────────────────────────────────────────────────────
def call_claude(prompt: str, label: str = '') -> list:
    client = anthropic.Anthropic(api_key=AKEY)
    for attempt in range(1, 4):
        try:
            msg = client.messages.create(
                model=MODEL, max_tokens=4096,
                messages=[{'role': 'user', 'content': prompt + '\n\nJSON만 출력: {"signals":[...]}'}])
            raw = msg.content[0].text.strip()
            for pat in [r'```json\s*(.*?)\s*```', r'(\{.*\})', None]:
                try:
                    text = re.search(pat, raw, re.DOTALL).group(1) if pat else raw
                    sigs = json.loads(text).get('signals', [])
                    return [s for s in sigs if isinstance(s, dict) and s.get('signal_type') in VALID]
                except:
                    pass
            log(f'  [JSON파싱실패] {label}')
            return []
        except anthropic.RateLimitError:
            log(f'  [RATE] 30초 대기')
            time.sleep(30)
        except Exception as e:
            log(f'  [ERR] {label}: {e}')
            if attempt < 3:
                time.sleep(5)
    return []

# ── 해당 stock 시그널 찾기 ────────────────────────────────────────────────────
def find_signal(signals: list, target_stock: str, target_ticker: str) -> str:
    """signals에서 target_stock/ticker에 해당하는 signal_type 반환. 없으면 '미탐지'"""
    target_stock  = target_stock.strip().lower()
    target_ticker = target_ticker.strip().upper()

    for s in signals:
        stock  = str(s.get('stock', '')).strip().lower()
        ticker = str(s.get('ticker', '')).strip().upper()
        if stock == target_stock or (target_ticker and ticker == target_ticker):
            return s.get('signal_type', '미탐지')

    # 부분 매칭 (종목명 포함 여부)
    for s in signals:
        stock = str(s.get('stock', '')).strip().lower()
        if target_stock in stock or stock in target_stock:
            return s.get('signal_type', '미탐지')

    return '미탐지'

# ── 엑셀 저장 ─────────────────────────────────────────────────────────────────
CLR = {
    'match':   PatternFill('solid', fgColor='C6EFCE'),
    'miss':    PatternFill('solid', fgColor='FFC7CE'),
    'header':  PatternFill('solid', fgColor='1F4E79'),
    'even':    PatternFill('solid', fgColor='F2F2F2'),
    'odd':     PatternFill('solid', fgColor='FFFFFF'),
    '매수':    PatternFill('solid', fgColor='DAEEF3'),
    '긍정':    PatternFill('solid', fgColor='EBF1DE'),
    '중립':    PatternFill('solid', fgColor='F2F2F2'),
    '부정':    PatternFill('solid', fgColor='FDE9D9'),
    '매도':    PatternFill('solid', fgColor='FFC7CE'),
    '미탐지':  PatternFill('solid', fgColor='FFFFCC'),
}
THIN = Border(*[Side(style='thin')]*0,
              left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

def xc(ws, r, c, val, fill=None, bold=False, align='center', color='000000'):
    cell = ws.cell(row=r, column=c, value=val)
    if fill: cell.fill = fill
    cell.font = Font(bold=bold, color=color, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = THIN
    return cell

def save_excel(rows, totals):
    wb = Workbook()
    ws = wb.active
    ws.title = '평가결과'

    headers = ['no','stock','ticker','jay_answer','V11.5','V12.0','V12.1',
               'V11.5 일치','V12.0 일치','V12.1 일치','video_id']
    widths  = [5, 16, 8, 10, 10, 10, 10, 10, 10, 10, 14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        xc(ws, 1, i, h, CLR['header'], bold=True, color='FFFFFF')
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    for ri, row in enumerate(rows, 2):
        rf = CLR['even'] if ri % 2 == 0 else CLR['odd']
        xc(ws, ri, 1,  row['no'],         rf, align='center')
        xc(ws, ri, 2,  row['stock'],       rf)
        xc(ws, ri, 3,  row['ticker'],      rf, align='center')
        xc(ws, ri, 4,  row['jay_answer'],  CLR.get(row['jay_answer'], rf), align='center')
        for ci, ver in enumerate(['v115','v120','v121'], 5):
            sig = row[ver]
            xc(ws, ri, ci, sig, CLR.get(sig, rf), align='center')
        for ci, ver in enumerate(['v115_match','v120_match','v121_match'], 8):
            m = row[ver]
            xc(ws, ri, ci, m, CLR['match'] if m == '✅' else CLR['miss'] if m == '❌' else rf, align='center')
        xc(ws, ri, 11, row['video_id'], rf)
        ws.row_dimensions[ri].height = 18

    # 합산 행
    last = len(rows) + 2
    xc(ws, last, 1, '합계', CLR['header'], bold=True, color='FFFFFF')
    xc(ws, last, 2, f'총 {len(rows)}개', CLR['header'], color='FFFFFF')
    xc(ws, last, 3, '', CLR['header'], color='FFFFFF')
    xc(ws, last, 4, '', CLR['header'], color='FFFFFF')
    for ci, ver in enumerate(['v115','v120','v121'], 5):
        xc(ws, last, ci, '', CLR['header'], color='FFFFFF')
    for ci, (ver, cnt) in enumerate(zip(['v115','v120','v121'], totals), 8):
        acc = f'{cnt}/{len(rows)} ({cnt/len(rows):.1%})'
        xc(ws, last, ci, acc, CLR['match'], bold=True, align='center')
    xc(ws, last, 11, '', CLR['header'], color='FFFFFF')

    ws.freeze_panes = 'A2'
    wb.save(RESULT_XLSX)
    log(f'엑셀 저장: {RESULT_XLSX}')

# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    log(f'=== eval_yt_v11v12.py | {MODEL} ===')

    # CSV 읽기
    csv_path = ROOT / 'data' / 'eval_ground_truth_40.csv'
    gt_rows = []
    with open(csv_path, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            gt_rows.append(row)
    log(f'GT 행 수: {len(gt_rows)}개')

    # 고유 video_id 추출
    unique_videos = {}
    for row in gt_rows:
        vid = row['video_id']
        if vid not in unique_videos:
            unique_videos[vid] = row['video_url']
    log(f'고유 영상: {len(unique_videos)}개\n')

    # 1단계: 자막 추출
    subtitles = {}
    log('── 1단계: 자막 추출 ──')
    for i, (vid, url) in enumerate(unique_videos.items(), 1):
        log(f'[{i}/{len(unique_videos)}] {vid}')
        sub = get_subtitle(vid)
        subtitles[vid] = sub
        if not sub:
            log(f'  ⚠ 자막 없음')
        time.sleep(1)

    # 2단계: 3개 버전 분석 (영상 단위)
    log('\n── 2단계: 프롬프트 분석 ──')
    analysis = {}  # video_id -> {v115: [...], v120: [...], v121: [...]}

    prompts = [
        ('v115', PROMPT_V115),
        ('v120', PROMPT_V120),
        ('v121', PROMPT_V121),
    ]

    for i, (vid, sub) in enumerate(subtitles.items(), 1):
        if not sub:
            analysis[vid] = {'v115': [], 'v120': [], 'v121': []}
            log(f'[{i}/{len(subtitles)}] {vid} — 자막 없음, 건너뜀')
            continue

        log(f'\n[{i}/{len(subtitles)}] {vid}')

        # V11.5 / V12.0 / V12.1 병렬 분석
        def _call(ver_ppath):
            ver, ppath = ver_ppath
            t0 = time.time()
            sigs = call_claude(build_prompt(ppath, sub, vid), f'{ver}/{vid}')
            log(f'  [{ver}] {time.time()-t0:.0f}초 | {len(sigs)}건: {[s.get("stock") for s in sigs]}')
            return ver, sigs

        with ThreadPoolExecutor(max_workers=3) as ex:
            futures = {ex.submit(_call, vp): vp[0] for vp in prompts}
            analysis[vid] = {}
            for fut in as_completed(futures):
                ver, sigs = fut.result()
                analysis[vid][ver] = sigs

        # 진행상황 저장
        with open(PROGRESS, 'w', encoding='utf-8') as f:
            json.dump({'subtitles_done': i, 'analysis': {
                k: {vv: vvv for vv, vvv in vd.items()} for k, vd in analysis.items()
            }}, f, ensure_ascii=False, indent=2)

        if i % 5 == 0 and i < len(subtitles):
            log('  [배치] 10초 대기')
            time.sleep(10)

    # 3단계: 결과 표 생성
    log('\n── 3단계: 결과 표 ──')
    result_rows = []
    cnt = {'v115': 0, 'v120': 0, 'v121': 0}

    for row in gt_rows:
        vid    = row['video_id']
        stock  = row['stock']
        ticker = row['ticker']
        jay    = row['jay_answer']
        an     = analysis.get(vid, {'v115': [], 'v120': [], 'v121': []})

        v115 = find_signal(an['v115'], stock, ticker)
        v120 = find_signal(an['v120'], stock, ticker)
        v121 = find_signal(an['v121'], stock, ticker)

        def match(pred): return '✅' if pred == jay else ('❌' if pred != '미탐지' else '❌')

        m115 = match(v115)
        m120 = match(v120)
        m121 = match(v121)

        if v115 == jay: cnt['v115'] += 1
        if v120 == jay: cnt['v120'] += 1
        if v121 == jay: cnt['v121'] += 1

        result_rows.append({
            'no': row['no'], 'video_id': vid, 'stock': stock, 'ticker': ticker,
            'jay_answer': jay,
            'v115': v115, 'v120': v120, 'v121': v121,
            'v115_match': m115, 'v120_match': m120, 'v121_match': m121,
        })

    total = len(result_rows)

    # 마크다운 출력
    md = [f'# V11.5 vs V12.0 vs V12.1 Eval ({TS})',
          f'모델: {MODEL} | 영상: {len(unique_videos)}개 | 평가: {total}행\n',
          f'| no | stock | jay_answer | v11.5 | v12.0 | v12.1 | v11.5_match | v12.0_match | v12.1_match |',
          f'|---|---|---|---|---|---|---|---|---|']

    for r in result_rows:
        md.append(f'| {r["no"]} | {r["stock"]} | {r["jay_answer"]} | {r["v115"]} | {r["v120"]} | {r["v121"]} | {r["v115_match"]} | {r["v120_match"]} | {r["v121_match"]} |')

    md.append(f'\n**합산:** V11.5={cnt["v115"]}/{total} ({cnt["v115"]/total:.1%}) | V12.0={cnt["v120"]}/{total} ({cnt["v120"]/total:.1%}) | V12.1={cnt["v121"]}/{total} ({cnt["v121"]/total:.1%})')

    md_text = '\n'.join(md)
    print('\n' + md_text)
    RESULT_MD.write_text(md_text, encoding='utf-8')
    log(f'\nMD 저장: {RESULT_MD}')

    save_excel(result_rows, [cnt['v115'], cnt['v120'], cnt['v121']])
    log('완료!')

if __name__ == '__main__':
    main()
